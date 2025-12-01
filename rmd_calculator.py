#!/usr/bin/env python3

import csv
from dataclasses import dataclass
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, numbers

# IRS Uniform Lifetime Table (Table III)
UNIFORM_LIFETIME_FACTORS = {
    73: 26.5, 74: 25.5, 75: 24.6, 76: 23.7, 77: 22.9,
    78: 22.0, 79: 21.1, 80: 20.2, 81: 19.4, 82: 18.5,
    83: 17.7, 84: 16.8, 85: 16.0, 86: 15.2, 87: 14.4,
    88: 13.7, 89: 12.9, 90: 12.2, 91: 11.5, 92: 10.8,
    93: 10.1, 94: 9.5, 95: 8.9, 96: 8.4, 97: 7.8,
    98: 7.3, 99: 6.8, 100: 6.4, 101: 6.0, 102: 5.6,
    103: 5.2, 104: 4.9, 105: 4.6, 106: 4.3, 107: 4.1,
    108: 3.9, 109: 3.7, 110: 3.5, 111: 3.4, 112: 3.3,
    113: 3.1, 114: 3.0, 115: 2.9, 116: 2.8, 117: 2.7,
    118: 2.5, 119: 2.3, 120: 2.0
}


def get_divisor(age: int) -> float:
    return UNIFORM_LIFETIME_FACTORS.get(
        age,
        UNIFORM_LIFETIME_FACTORS[max(UNIFORM_LIFETIME_FACTORS.keys())]
    )


def calculate_rmd(balance: float, age: int) -> float:
    divisor = get_divisor(age)
    return balance / divisor


def run_projection(start_age, start_balance, years, growth_rate, withholding_rate):
    growth = growth_rate / 100.0
    tax_rate = withholding_rate / 100.0

    age = start_age
    balance = start_balance

    rows = []

    for year in range(years):
        if age < 73:
            rmd = 0
        else:
            rmd = calculate_rmd(balance, age)

        tax_withheld = rmd * tax_rate
        net_to_you = rmd - tax_withheld

        end_balance = (balance - rmd) * (1 + growth)

        rows.append({
            "Year": year + 1,
            "Age": age,
            "Start Balance": balance,
            "RMD": rmd,
            "Tax Withheld": tax_withheld,
            "Net Received": net_to_you,
            "End Balance": end_balance,
        })

        balance = end_balance
        age += 1

    return rows


def export_xlsx(filename, rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "RMD Projection"

    headers = list(rows[0].keys())

    # Write header row
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    # Write data rows
    for row_idx, row_data in enumerate(rows, start=2):
        for col_idx, key in enumerate(headers, start=1):
            value = row_data[key]
            cell = ws.cell(row=row_idx, column=col_idx, value=value)

            # Currency formatting for money values
            if key in ("Start Balance", "RMD", "Tax Withheld", "Net Received", "End Balance"):
                cell.number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE

    # Auto column width
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                length = len(str(cell.value))
                max_length = max(max_length, length)
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2

    wb.save(filename)
    print(f"\nXLSX file saved as: {filename}")


def prompt_float(msg):
    return float(input(msg).replace(",", "").strip())


def prompt_int(msg):
    return int(input(msg).strip())


def main():
    print("=== IRA RMD Calculator + Projection + XLSX Export ===\n")

    age = prompt_int("Your age this year: ")
    balance = prompt_float("IRA balance as of last Dec 31: $")
    years = prompt_int("Project how many years? ")
    growth_rate = prompt_float("Annual growth rate (e.g., 5 for 5%): ")
    withholding = prompt_float("Tax withholding percentage (e.g., 20): ")

    rows = run_projection(age, balance, years, growth_rate, withholding)

    print("\n--- Projection ---")
    print(f"{'Yr':<3} {'Age':<4} {'Start Balance':>15} {'RMD':>12} "
          f"{'Tax':>12} {'Net Rcvd':>12} {'End Balance':>15}")

    for r in rows:
        print(f"{r['Year']:<3} {r['Age']:<4} "
              f"${r['Start Balance']:>14,.2f} "
              f"${r['RMD']:>11,.2f} "
              f"${r['Tax Withheld']:>11,.2f} "
              f"${r['Net Received']:>11,.2f} "
              f"${r['End Balance']:>14,.2f}")

    # === XLSX Export Option ===
    save_xlsx = input("\nExport results to XLSX? (y/n): ").strip().lower()
    if save_xlsx == "y":
        filename = input("Enter XLSX filename (e.g., rmd_projection.xlsx): ").strip()
        if not filename.lower().endswith(".xlsx"):
            filename += ".xlsx"
        export_xlsx(filename, rows)

    print("\nDone.")


if __name__ == "__main__":
    main()
